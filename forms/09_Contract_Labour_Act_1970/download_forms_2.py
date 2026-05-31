"""
Download Statutory Labour Law Forms — with Full Result Reporting
================================================================
Run on YOUR PC:
    pip install requests openpyxl
    python download_forms.py

After running you will see:
  • Live progress on screen (each form: ✓ or ✗)
  • Final summary table in console
  • Excel log: E:\\AICA 2\\LABOUR LAW COMPILER\\forms\\DOWNLOAD_REPORT.xlsx
      - Sheet 1: Summary (counts)
      - Sheet 2: All forms (status, file size, path, URL used)
      - Sheet 3: Failed only (with browser URLs to manually download)
"""

import os, re, sys, time
from pathlib import Path
from datetime import datetime
import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False
    print("NOTE: openpyxl not installed — Excel report will be skipped.")
    print("      Run: pip install openpyxl\n")

# ─── CONFIG ───────────────────────────────────────────────────────────────────
BASE_FOLDER = Path(r"E:\AICA 2\LABOUR LAW COMPILER\forms")
DELAY       = 2       # seconds between downloads
TIMEOUT     = 60      # seconds per request

HEADERS = {
    "User-Agent"     : "Mozilla/5.0 (Windows NT 10.0; Win64; x64) Chrome/124.0.0.0 Safari/537.36",
    "Accept"         : "application/pdf,application/octet-stream,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

# ─── FORMS LIST ───────────────────────────────────────────────────────────────
# (subfolder, filename, description, [url1, url2, url3...])
# URLs tried in order — first success wins.

FORMS = [

    # ── PAYMENT OF BONUS ACT, 1965 ────────────────────────────────────────────
    ("01_Bonus_Act_1965",
     "Bonus_Rules_1975_Forms_ABCD.pdf",
     "Forms A B C D — Bonus Rules 1975 (all 4 forms inside)",
     ["https://clc.gov.in/clc/sites/default/files/BonusAct.pdf",
      "https://labour.gov.in/sites/default/files/thepaymentof_bonus_rules1975.pdf",
      "https://www.labour.gov.in/static/uploads/2025/06/3fbc98da530fbc248c7a0f711dc5dd2b.pdf"]),

    # ── MATERNITY BENEFIT ACT, 1961 ───────────────────────────────────────────
    ("02_Maternity_Benefit_Act_1961",
     "Maternity_Benefit_Rules_1963_Form_A_Muster_Roll.pdf",
     "Form A — Muster Roll (Maternity Benefit Rules, 1963)",
     ["https://clc.gov.in/clc/sites/default/files/MaternityBenefitAct.pdf",
      "https://labour.gov.in/sites/default/files/maternity_benefit_rules1963.pdf"]),

    # ── PAYMENT OF GRATUITY ACT, 1972 ────────────────────────────────────────
    ("03_Gratuity_Act_1972",
     "Gratuity_Act_Central_Rules_1972_All_Forms.pdf",
     "Forms A B C F + all — Gratuity Act & Central Rules",
     ["https://clc.gov.in/clc/sites/default/files/PaymentofGratuityAct.pdf",
      "https://labour.gov.in/sites/default/files/payment_of_gratuity_central_rules_1972.pdf"]),

    ("03_Gratuity_Act_1972",
     "Form_T_Recovery_of_Gratuity.pdf",
     "Form T — Application for Recovery of Gratuity",
     ["https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/Form%20T.pdf",
      "https://clc.gov.in/clc/sites/default/files/FormT_Gratuity.pdf"]),

    # ── MINIMUM WAGES ACT, 1948 ───────────────────────────────────────────────
    ("04_Minimum_Wages_Act_1948",
     "Minimum_Wages_Central_Rules_1950_Forms_I_to_IV.pdf",
     "Forms I II III IV — Minimum Wages Central Rules, 1950",
     ["https://clc.gov.in/clc/sites/default/files/MinimumWagesact.pdf",
      "https://labour.gov.in/sites/default/files/minimum_wages_central_rules1950.pdf"]),

    ("04_Minimum_Wages_Act_1948",
     "Form_VI_Application_Employee_MinWages.pdf",
     "Form VI — Application by Employee under Sec 20(2) Min Wages Act",
     ["https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/FORM%20VI.pdf",
      "https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/Form%20VI.pdf"]),

    ("04_Minimum_Wages_Act_1948",
     "Form_VIA_Group_Application_MinWages.pdf",
     "Form VI-A — Group Application under Sec 21(1) Min Wages Act",
     ["https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/FORM%20VI-A.pdf",
      "https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/Form%20VI-A.pdf"]),

    # ── ESI ACT, 1948 ─────────────────────────────────────────────────────────
    ("05_ESI_Act_1948",
     "ESI_General_Regulations_1950_All_Forms.pdf",
     "All ESI Forms — ESI General Regulations 1950",
     ["https://esic.gov.in/attachments/regularionact/ESI_General_Regulations_1950.pdf",
      "https://esic.gov.in/attachments/regularionact/ESIGeneralRegulations1950.pdf",
      "https://esic.gov.in/sites/default/files/ESI_Regulations.pdf"]),

    ("05_ESI_Act_1948",
     "Form_1_Employer_Declaration.pdf",
     "Form 1 — Declaration / Employer Registration (ESI Act)",
     ["https://esic.gov.in/attachments/regularionact/form1.pdf",
      "https://esic.gov.in/sites/default/files/form1.pdf"]),

    ("05_ESI_Act_1948",
     "Form_3_Return_of_Declaration.pdf",
     "Form 3 — Return of Declaration (ESI Act)",
     ["https://esic.gov.in/attachments/regularionact/form3.pdf",
      "https://esic.gov.in/sites/default/files/form3.pdf"]),

    ("05_ESI_Act_1948",
     "Form_5_Return_of_Contributions.pdf",
     "Form 5 — Return of Contributions (ESI Act)",
     ["https://esic.gov.in/attachments/regularionact/form5.pdf",
      "https://esic.gov.in/sites/default/files/form5.pdf"]),

    ("05_ESI_Act_1948",
     "Form_6_Register_of_Employees.pdf",
     "Form 6 — Register of Employees (ESI Act)",
     ["https://esic.gov.in/attachments/regularionact/form6.pdf",
      "https://esic.gov.in/sites/default/files/form6.pdf"]),

    # ── EPF ACT, 1952 ─────────────────────────────────────────────────────────
    ("06_EPF_Act_1952",
     "Form_2_Nomination_Declaration.pdf",
     "Form 2 — Nomination & Declaration (EPF)",
     ["https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/Form2Revised.pdf",
      "https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/Form2.pdf"]),

    ("06_EPF_Act_1952",
     "Form_11_New_Employee_Declaration.pdf",
     "Form 11 — Declaration by New Employee (EPF)",
     ["https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/Form11Revised.pdf",
      "https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/Form11_New.pdf",
      "https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/Form11.pdf"]),

    ("06_EPF_Act_1952",
     "Form_12A_Annual_Return.pdf",
     "Form 12A — Annual Return (EPF)",
     ["https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/Form12A.pdf",
      "https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/FORM_12A.pdf"]),

    ("06_EPF_Act_1952",
     "Form_19_PF_Final_Settlement.pdf",
     "Form 19 — PF Final Settlement / Withdrawal (EPF)",
     ["https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/Form19UAN.pdf",
      "https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/FORM_19.pdf",
      "https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/Form19.pdf"]),

    ("06_EPF_Act_1952",
     "Form_10C_EPS_Withdrawal.pdf",
     "Form 10C — EPS Withdrawal / Scheme Certificate (EPF)",
     ["https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/Form10C.pdf",
      "https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/FORM_10C.pdf"]),

    ("06_EPF_Act_1952",
     "Form_10D_Monthly_Pension.pdf",
     "Form 10D — Monthly Pension (EPS 1995)",
     ["https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/Form10D.pdf",
      "https://www.epfindia.gov.in/site_docs/PDFs/Misc_PDFs/FORM_10D.pdf"]),

    # ── FACTORIES ACT, 1948 ───────────────────────────────────────────────────
    ("07_Factories_Act_1948",
     "Factories_Central_Rules_All_Forms_25_26.pdf",
     "Forms 25 26 16 + all — Factories Act Central Rules",
     ["https://clc.gov.in/clc/sites/default/files/FactoriesAct.pdf",
      "https://labour.gov.in/sites/default/files/factories_central_rules1950.pdf"]),

    # ── INDUSTRIAL DISPUTES ACT, 1947 ─────────────────────────────────────────
    ("08_Industrial_Disputes_Act_1947",
     "Industrial_Disputes_Central_Rules_Form_A.pdf",
     "Form A Notice of Change + all — ID Central Rules 1957",
     ["https://clc.gov.in/clc/sites/default/files/IndustrialDisputesAct.pdf",
      "https://labour.gov.in/sites/default/files/industrial_disputes_central_rules1957.pdf"]),

    ("08_Industrial_Disputes_Act_1947",
     "Format_Filing_Industrial_Dispute_CLC.pdf",
     "Format for Filing Industrial Dispute (CLC New Delhi)",
     ["https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/Format%20for%20filing%20Industrial%20Dispute.pdf"]),

    # ── CONTRACT LABOUR ACT, 1970 ─────────────────────────────────────────────
    ("09_Contract_Labour_Act_1970",
     "Contract_Labour_Central_Rules_1971_All_Forms.pdf",
     "All Forms I IV VII XII XIII — Contract Labour Central Rules 1971",
     ["https://clc.gov.in/clc/sites/default/files/ContractLabourAct.pdf",
      "https://labour.gov.in/sites/default/files/contract_labour_central_rules1971.pdf"]),

    ("09_Contract_Labour_Act_1970",
     "Form_II_Application_Licence_Contractor.pdf",
     "Form II — Application for Licence/Renewal (Contractor)",
     ["https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/FORM%20II.pdf",
      "https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/Form%20II.pdf"]),

    ("09_Contract_Labour_Act_1970",
     "Form_III_Certificate_Principal_Employer.pdf",
     "Form III — Certificate by Principal Employer",
     ["https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/FORM%20III.pdf",
      "https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/Form%20III.pdf"]),

    # ── CHILD LABOUR ACT, 1986 ────────────────────────────────────────────────
    ("10_Child_Labour_Act_1986",
     "Child_Labour_Rules_1988_Form_01A.pdf",
     "Form 01-A — Annual Return (Child Labour Rules, 1988)",
     ["https://clc.gov.in/clc/sites/default/files/ChildLabourAct.pdf",
      "https://labour.gov.in/sites/default/files/child_labour_rules1988.pdf"]),

    # ── BOCW ACT, 1996 ────────────────────────────────────────────────────────
    ("11_BOCW_Act_1996",
     "BOCW_Central_Rules_1998_All_Forms.pdf",
     "All BOCW Forms I XXII + — BOCW Central Rules 1998",
     ["https://clc.gov.in/clc/sites/default/files/BOCWAct.pdf",
      "https://labour.gov.in/sites/default/files/BOCW_Central_Rules_1998.pdf"]),

    # ── POSH ACT, 2013 ────────────────────────────────────────────────────────
    ("12_POSH_Act_2013",
     "POSH_Act_2013_Rules_ICC_Complaint_Form.pdf",
     "POSH Act 2013 + Rules — ICC Complaint Form & all forms",
     ["https://www.indiacode.nic.in/bitstream/123456789/15815/1/posh_act_2013.pdf",
      "https://labour.gov.in/sites/default/files/posh_act_2013.pdf",
      "https://wcd.nic.in/sites/default/files/SHact.pdf"]),

    # ── PAYMENT OF WAGES ACT, 1936 ────────────────────────────────────────────
    ("13_Payment_of_Wages_Act_1936",
     "Payment_of_Wages_Act_Rules_All_Forms.pdf",
     "All Forms — Payment of Wages Act & Rules 1937",
     ["https://clc.gov.in/clc/sites/default/files/PaymentofWagesAct.pdf",
      "https://labour.gov.in/sites/default/files/payment_of_wages_rules1937.pdf"]),

    ("13_Payment_of_Wages_Act_1936",
     "Form_A_Individual_Application.pdf",
     "Form A — Individual Application under Sec 15(2) Payment of Wages Act",
     ["https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/FORM%20A.pdf",
      "https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/Form%20A.pdf"]),

    ("13_Payment_of_Wages_Act_1936",
     "Form_B_Group_Application.pdf",
     "Form B — Group Application under Sec 15(2) Payment of Wages Act",
     ["https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/FORM%20B.pdf",
      "https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/Form%20B.pdf"]),

    # ── GENERAL / OTHERS ──────────────────────────────────────────────────────
    ("14_General_Formats",
     "Equal_Remuneration_Rules_1976_Forms_DE.pdf",
     "Forms D E — Equal Remuneration Rules 1976",
     ["https://clc.gov.in/clc/sites/default/files/EqualRemunerationAct.pdf",
      "https://labour.gov.in/sites/default/files/equal_remuneration_rules1976.pdf"]),

    ("14_General_Formats",
     "Model_Standing_Orders_IE_Standing_Orders_Act.pdf",
     "Model Standing Orders — IE (Standing Orders) Act 1946",
     ["https://clc.gov.in/clc/sites/default/files/IndustrialEmploymentAct.pdf",
      "https://www.indiacode.nic.in/bitstream/123456789/1383/1/standing_orders_act_1946.pdf"]),

    ("14_General_Formats",
     "Interstate_Migrant_Workmen_Central_Rules_1980.pdf",
     "All Forms — Interstate Migrant Workmen Central Rules 1980",
     ["https://clc.gov.in/clc/sites/default/files/ISMWAct.pdf",
      "https://labour.gov.in/sites/default/files/interstate_migrant_workmen_central_rules1980.pdf"]),

    ("14_General_Formats",
     "Format_Short_Payment_Complaint_CLC.pdf",
     "Format — Short Payment Complaint (CLC)",
     ["https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/Format%20for%20filing%20Short%20Payment%20Complaints.pdf"]),

    ("14_General_Formats",
     "Format_Non_Payment_Complaint_CLC.pdf",
     "Format — Non Payment Complaint (CLC)",
     ["https://clc.gov.in/clc/Regionaloffice/NewDelhi/sites/default/files/Format%20for%20filing%20Non%20Payment%20Complaints.pdf"]),
]

TOTAL = len(FORMS)

# ─── DOWNLOAD ENGINE ──────────────────────────────────────────────────────────

def try_download(urls, dest):
    """Try each URL in order. Returns (success, url_used, size_kb, error_msg)."""
    if dest.exists() and dest.stat().st_size > 3000:
        kb = dest.stat().st_size // 1024
        return True, "already_existed", kb, ""

    dest.parent.mkdir(parents=True, exist_ok=True)
    last_error = ""

    for url in urls:
        try:
            h = {**HEADERS, "Referer": "/".join(url.split("/")[:3]) + "/"}
            r = requests.get(url, headers=h, timeout=TIMEOUT,
                             stream=True, allow_redirects=True)

            if r.status_code != 200:
                last_error = f"HTTP {r.status_code}"
                continue

            ct = r.headers.get("Content-Type", "").lower()
            if "html" in ct and "pdf" not in ct:
                last_error = f"Got HTML not PDF (CT={ct[:30]})"
                continue

            with open(dest, "wb") as f:
                for chunk in r.iter_content(16384):
                    if chunk:
                        f.write(chunk)

            sz = dest.stat().st_size
            if sz < 3000:
                dest.unlink()
                last_error = f"File too small ({sz} bytes)"
                continue

            return True, url, sz // 1024, ""

        except Exception as e:
            last_error = str(e)[:80]
            if dest.exists():
                dest.unlink()

    return False, "", 0, last_error


# ─── EXCEL REPORT ─────────────────────────────────────────────────────────────

def write_excel(results, report_path):
    if not HAS_EXCEL:
        return

    wb = openpyxl.Workbook()

    # colour helpers
    def hdr(ws, row, cols, bg):
        fill = PatternFill("solid", fgColor=bg)
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22

    def autowidth(ws, max_w=80):
        for col in ws.columns:
            ml = max((len(str(c.value)) if c.value else 0) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+3, max_w)

    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_rows(ws, start=2):
        for ri, row in enumerate(ws.iter_rows(min_row=start), start):
            bg = "F7F7F7" if ri % 2 == 0 else "FFFFFF"
            for cell in row:
                cell.fill      = PatternFill("solid", fgColor=bg)
                cell.border    = bdr
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    ok      = [r for r in results if r["status"] == "✓ Downloaded"]
    skipped = [r for r in results if r["status"] == "⏭ Already existed"]
    failed  = [r for r in results if r["status"] == "✗ Failed"]

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    hdr(ws1, 1, 2, "4472C4")
    ws1["A1"] = "Category"
    ws1["B1"] = "Count"

    summary_data = [
        ("Total Forms",        len(results)),
        ("✓ Downloaded",       len(ok)),
        ("⏭ Already Existed",  len(skipped)),
        ("✗ Failed",           len(failed)),
        ("",                   ""),
        ("Run Timestamp",      datetime.now().strftime("%d-%b-%Y %H:%M")),
        ("Save Folder",        str(BASE_FOLDER)),
    ]
    colors = {"✓": "EAF3DE", "⏭": "FFF9E6", "✗": "FDECEA"}
    for ri, (cat, val) in enumerate(summary_data, 2):
        ws1.cell(ri, 1).value = cat
        ws1.cell(ri, 2).value = val
        bg = next((v for k, v in colors.items() if str(cat).startswith(k)), "FFFFFF")
        for c in range(1, 3):
            ws1.cell(ri, c).fill      = PatternFill("solid", fgColor=bg)
            ws1.cell(ri, c).border    = bdr
            ws1.cell(ri, c).alignment = Alignment(vertical="center")

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 55

    # ── Sheet 2: All Forms ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Forms")
    cols2 = ["#", "Status", "Act / Category", "Form Description",
             "Filename", "Size (KB)", "URL Used", "Local Path", "Error"]
    hdr(ws2, 1, len(cols2), "4472C4")
    for ci, col in enumerate(cols2, 1):
        ws2.cell(1, ci).value = col

    for ri, r in enumerate(results, 2):
        row_bg = "EAF3DE" if "Downloaded" in r["status"] else \
                 "FFF9E6" if "existed"    in r["status"] else "FDECEA"
        vals = [r["sr"], r["status"], r["folder"], r["desc"],
                r["filename"], r["size_kb"], r["url_used"],
                r["local_path"], r["error"]]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(ri, ci)
            cell.value     = v
            cell.fill      = PatternFill("solid", fgColor=row_bg)
            cell.border    = bdr
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws2.freeze_panes = "A2"
    autowidth(ws2)

    # ── Sheet 3: Failed Only ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("❌ Failed — Manual Download")
    if failed:
        cols3 = ["#", "Act / Category", "Form Description",
                 "Filename", "Error", "Try URL 1", "Try URL 2", "Save To Folder"]
        hdr(ws3, 1, len(cols3), "C00000")
        for ci, col in enumerate(cols3, 1):
            ws3.cell(1, ci).value = col

        for ri, r in enumerate(failed, 2):
            urls = r["all_urls"]
            vals = [r["sr"], r["folder"], r["desc"], r["filename"],
                    r["error"],
                    urls[0] if len(urls) > 0 else "",
                    urls[1] if len(urls) > 1 else "",
                    str(BASE_FOLDER / r["folder"])]
            for ci, v in enumerate(vals, 1):
                cell = ws3.cell(ri, ci)
                cell.value     = v
                cell.fill      = PatternFill("solid", fgColor="FDECEA" if ri%2==0 else "FFF5F5")
                cell.border    = bdr
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        ws3.freeze_panes = "A2"
        autowidth(ws3)
        ws3["A1"].value = f"# Failed: {len(failed)} — Open each URL in browser and save to the specified folder"
    else:
        ws3["A1"].value = "🎉 All forms downloaded successfully — nothing failed!"
        ws3["A1"].font  = Font(bold=True, color="3B6D11", size=13)

    wb.save(report_path)


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def run():
    BASE_FOLDER.mkdir(parents=True, exist_ok=True)
    for sf in set(f[0] for f in FORMS):
        (BASE_FOLDER / sf).mkdir(parents=True, exist_ok=True)

    results  = []
    cur_fld  = ""
    t_start  = datetime.now()

    # column widths for console table
    W_DESC = 60
    W_STAT = 18
    W_SIZE =  8

    print()
    print("╔" + "═"*90 + "╗")
    print("║" + "  STATUTORY LABOUR LAW FORMS DOWNLOADER".center(90) + "║")
    print("║" + f"  {TOTAL} forms  →  {BASE_FOLDER}".center(90) + "║")
    print("╚" + "═"*90 + "╝")
    print()

    for i, (subfolder, filename, desc, urls) in enumerate(FORMS, 1):

        # section header
        if subfolder != cur_fld:
            cur_fld = subfolder
            label   = subfolder.replace("_", " ").strip()
            print(f"\n  {'─'*86}")
            print(f"  {label}")
            print(f"  {'─'*86}")

        dest = BASE_FOLDER / subfolder / filename

        # live progress line
        short_desc = desc[:W_DESC]
        sys.stdout.write(f"  [{i:02d}/{TOTAL}]  {short_desc:<{W_DESC}}  ...")
        sys.stdout.flush()

        ok, url_used, size_kb, err = try_download(urls, dest)

        if ok:
            if url_used == "already_existed":
                status = "⏭  Already existed"
                row_status = "⏭ Already existed"
            else:
                status = f"✓  Downloaded"
                row_status = "✓ Downloaded"
        else:
            status = "✗  FAILED"
            row_status = "✗ Failed"

        size_str = f"{size_kb} KB" if size_kb else "—"
        print(f"\r  [{i:02d}/{TOTAL}]  {short_desc:<{W_DESC}}  {status:<{W_STAT}}  {size_str:>{W_SIZE}}")

        results.append({
            "sr"         : i,
            "status"     : row_status,
            "folder"     : subfolder.replace("_"," ").strip(),
            "desc"       : desc,
            "filename"   : filename,
            "size_kb"    : size_kb if ok else 0,
            "url_used"   : url_used if ok else "",
            "local_path" : str(dest) if ok else "",
            "error"      : err,
            "all_urls"   : urls,
        })

        time.sleep(DELAY)

    # ── Console summary ───────────────────────────────────────────────────────
    ok_n  = sum(1 for r in results if "Downloaded" in r["status"])
    sk_n  = sum(1 for r in results if "existed"    in r["status"])
    fa_n  = sum(1 for r in results if "Failed"     in r["status"])
    secs  = (datetime.now() - t_start).seconds

    print()
    print("  " + "═"*88)
    print(f"  {'FINAL RESULT':}")
    print("  " + "═"*88)
    print(f"  {'Total forms':<30} {TOTAL}")
    print(f"  {'✓ Downloaded':<30} {ok_n}")
    print(f"  {'⏭ Already existed (skipped)':<30} {sk_n}")
    print(f"  {'✗ Failed':<30} {fa_n}")
    print(f"  {'Time taken':<30} {secs}s")
    print("  " + "─"*88)

    if fa_n:
        print(f"\n  ✗ FAILED FORMS — open these URLs in your browser:\n")
        for r in results:
            if "Failed" not in r["status"]:
                continue
            print(f"  [{r['sr']:02d}] {r['desc']}")
            print(f"       Folder : {BASE_FOLDER / r['folder']}")
            for u in r["all_urls"]:
                print(f"       URL    : {u}")
            print()
    else:
        print("\n  🎉  All forms downloaded successfully!\n")

    # ── Excel report ──────────────────────────────────────────────────────────
    report_path = BASE_FOLDER / "DOWNLOAD_REPORT.xlsx"
    write_excel(results, report_path)
    if HAS_EXCEL:
        print(f"  📊  Excel report saved: {report_path}")
        print(f"       Sheet 1 = Summary | Sheet 2 = All Forms | Sheet 3 = Failed Only")

    print()


if __name__ == "__main__":
    run()
